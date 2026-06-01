import re
from pathlib import Path

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side


SOURCE = Path(
    "/Users/wonsik/Documents/3_1_season/소프트웨어공학/sw-project-v1/Requirement/"
    "SE-0.SW요구사항-강의실예약시스템-1조.xlsx"
)


def normalize(value):
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").strip()


def category_for(req_id):
    if req_id.startswith("SFR-0"):
        return "시스템 로그인"
    if req_id.startswith("SFR-1"):
        return "사용자 정보 관리"
    if req_id.startswith("SFR-2"):
        return "강의실 현황 조회"
    if req_id.startswith("SFR-3"):
        return "강의실 현황 관리"
    if req_id.startswith("SFR-4"):
        return "예약 생성"
    if req_id.startswith("SFR-5"):
        return "예약 상태 관리"
    if req_id.startswith("SFR-6"):
        return "데이터 관리"
    if req_id.startswith("SFR-7"):
        return "네트워크 및 세션 관리"
    if req_id.startswith("NFR-"):
        return "비기능 요구사항"
    return "기타"


def test_type_for(category, requirement):
    text = requirement
    if category == "비기능 요구사항":
        if "1초" in text or "3초" in text or "50명" in text:
            return "경계값 분석"
        if "MacOS" in text or "Window" in text or "JRE" in text:
            return "기타"
        if "개인정보" in text:
            return "기타"
        return "기타"
    if "로그인" in category:
        return "유스케이스 테스팅"
    if "사용자" in category:
        return "문장 테스팅"
    if "예약" in category:
        return "유스케이스 테스팅"
    if "네트워크" in category:
        return "기타"
    return "문장 테스팅"


def precondition_for(category, requirement):
    if category == "시스템 로그인":
        return "서버가 실행 중이고 User.json에 테스트 사용자 계정이 존재한다."
    if category == "사용자 정보 관리":
        return "관리자 계정으로 로그인되어 있고 사용자 관리 화면에 접근 가능하다."
    if category == "강의실 현황 조회":
        return "강의실 정보와 시간표/예약 데이터가 테스트 데이터로 등록되어 있다."
    if category == "강의실 현황 관리":
        return "조교 계정으로 로그인되어 있고 강의실 관리 화면에 접근 가능하다."
    if category == "예약 생성":
        return "교수 또는 학생 계정으로 로그인되어 있고 예약 가능한 강의실 데이터가 존재한다."
    if category == "예약 상태 관리":
        return "조교 계정으로 로그인되어 있고 승인 대기 예약 신청 데이터가 존재한다."
    if category == "데이터 관리":
        return "백업/복구 대상 JSON 데이터 파일이 준비되어 있다."
    if category == "네트워크 및 세션 관리":
        return "서버가 실행 중이고 테스트 클라이언트 연결을 생성할 수 있다."
    return "테스트에 필요한 기본 데이터가 준비되어 있다."


def data_for(req_id, category, requirement, note):
    if category == "시스템 로그인":
        return "관리자 admin/admin, 조교 23456/23456, 교수 34567/34567, 학생 20240001/20240001"
    if category == "사용자 정보 관리":
        return "추가 사용자: 권한=STUDENT, ID=20240001, 이름=테스트학생 / 삭제 대상: 23456"
    if category == "강의실 현황 조회":
        return "건물=정보관, 강의실=911, 기준일=테스트 날짜"
    if category == "강의실 현황 관리":
        return "강의실=911, 수용인원=40, 사용 가능 컴퓨터=40, 시간표 PDF=테스트 파일"
    if category == "예약 생성":
        return "강의실=911, 예약일=현재일+1~14일, 시간=09:00~12:00, 목적=보강/세미나/개인 학습"
    if category == "예약 상태 관리":
        return "예약 ID=R-TEST-001, 상태=PENDING, 거부 사유=수업 일정과 중복"
    if category == "데이터 관리":
        return "User.json, Classroom.json, Reservation.json 테스트 파일"
    if category == "네트워크 및 세션 관리":
        return "정상 JSON 요청, 비정상 JSON 문자열, heartbeat 미응답 클라이언트"
    if req_id.startswith("NFR-"):
        return normalize(note) or "성능/호환성/표준 검증용 테스트 데이터"
    return normalize(note)


def expected_for(requirement):
    text = normalize(requirement)
    if "성공/실패" in text:
        return "성공 또는 실패 여부가 사용자에게 명확한 메시지로 표시된다."
    if "실패" in text and "로그인 초기 화면" in text:
        return "로그인 실패 후 로그인 초기 화면으로 돌아간다."
    if "성공" in text and "권한에 맞는 화면" in text:
        return "로그인 성공 후 사용자 권한에 맞는 화면이 표시된다."
    if "추가" in text:
        return "입력한 데이터가 저장되고 목록 또는 파일에 반영된다."
    if "삭제" in text:
        return "선택한 데이터가 삭제되고 결과 메시지가 표시된다."
    if "조회" in text:
        return "요구사항에 명시된 정보가 누락 없이 조회되고 필요한 기준으로 정렬된다."
    if "승인" in text:
        return "선택한 예약의 상태가 승인으로 변경되고 관련 사용자에게 결과가 전달된다."
    if "거부" in text:
        return "선택한 예약의 상태가 취소/거부로 변경되고 거부 사유가 전달된다."
    if "제한" in text or "불가" in text:
        return "제한 조건을 위반한 입력은 거부되고 데이터가 변경되지 않는다."
    if "저장" in text or "유지" in text:
        return "프로그램 종료 또는 재시작 후에도 데이터가 유지된다."
    if "1초" in text or "3초" in text:
        return "요구사항에 명시된 시간 이내에 처리가 완료된다."
    return "실제 결과가 SRS 요구사항의 조건과 일치한다."


def test_name_for(req_id, category, requirement):
    text = requirement
    if category == "시스템 로그인":
        if "권한과 사용자 정보" in text:
            return "로그인-권한정보검증"
        if "성공/실패" in text:
            return "로그인-결과알림"
        if "실패" in text:
            return "로그인-실패흐름"
        if "성공" in text:
            return "로그인-성공흐름"
        if "로그아웃" in text:
            return "로그아웃-초기화면"
        return "로그인-기본흐름"
    if category == "사용자 정보 관리":
        if "추가" in text:
            return "사용자관리-사용자추가"
        if "삭제" in text:
            return "사용자관리-사용자삭제"
        return "사용자관리-기본기능"
    if category == "강의실 현황 조회":
        return "강의실현황조회"
    if category == "강의실 현황 관리":
        return "강의실현황관리"
    if category == "예약 생성":
        if "교수" in text:
            return "예약생성-교수"
        if "학생" in text:
            return "예약생성-학생"
        return "예약생성-공통"
    if category == "예약 상태 관리":
        if "승인" in text:
            return "예약상태관리-승인"
        if "거부" in text:
            return "예약상태관리-거부"
        if "취소" in text:
            return "예약상태관리-취소"
        return "예약상태관리"
    if category == "데이터 관리":
        return "데이터관리"
    if category == "네트워크 및 세션 관리":
        return "네트워크세션관리"
    return f"{req_id}-검증"


def related_class_for(category):
    if category == "시스템 로그인":
        return "LoginService, UserService, User"
    if category == "사용자 정보 관리":
        return "UserService, UserFileManager, UserCatalog, User"
    if category == "데이터 관리":
        return "UserFileManager"
    return ""


def load_requirements(wb):
    ws = wb["SRS"]
    requirements = []
    for row in range(3, ws.max_row + 1):
        req_id = normalize(ws.cell(row, 3).value)
        stakeholder = normalize(ws.cell(row, 4).value)
        requirement = normalize(ws.cell(row, 5).value)
        note = normalize(ws.cell(row, 6).value)

        if not req_id or not requirement:
            continue
        if not re.match(r"^(SFR|NFR)-\d{3}$", req_id):
            continue

        requirements.append((req_id, stakeholder, requirement, note))
    return requirements


def apply_styles(ws):
    header_fill = PatternFill("solid", fgColor="FFFF00")
    test_type_fill = PatternFill("solid", fgColor="F4B183")
    editable_fill = PatternFill("solid", fgColor="F4B6B6")
    header_font = Font(bold=True, size=12)
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    ws["E1"].fill = test_type_fill

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=11):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
        for cell in row[4:7]:
            cell.fill = editable_fill

    widths = {
        "A": 10,
        "B": 14,
        "C": 28,
        "D": 42,
        "E": 18,
        "F": 24,
        "G": 20,
        "H": 30,
        "I": 34,
        "J": 42,
        "K": 14,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 28
    for idx in range(2, ws.max_row + 1):
        ws.row_dimensions[idx].height = 72

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:K{ws.max_row}"

    test_types = "동등분할,경계값 분석,구문 테스팅,유스케이스 테스팅,문장 테스팅,결정/분기 테스팅,조건 테스팅,기타"
    validation = DataValidation(type="list", formula1=f'"{test_types}"', allow_blank=True)
    ws.add_data_validation(validation)
    validation.add(f"E2:E{ws.max_row}")


def main():
    wb = openpyxl.load_workbook(SOURCE)
    requirements = load_requirements(wb)

    sheet_name = "SRS Test Cases"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)
    headers = [
        "ID",
        "요구사항 ID",
        "테스트 케이스 이름",
        "설명",
        "테스트 유형",
        "관련 클래스",
        "관련 유스케이스",
        "초기 조건",
        "입력값",
        "예상 결과",
        "테스팅 결과",
    ]
    ws.append(headers)

    for index, (req_id, stakeholder, requirement, note) in enumerate(requirements, start=1):
        category = category_for(req_id)
        tc_id = f"TC-{index:02d}"
        description = requirement
        if stakeholder:
            description += f"\n대상: {stakeholder}"

        ws.append(
            [
                tc_id,
                req_id,
                test_name_for(req_id, category, requirement),
                description,
                test_type_for(category, requirement),
                related_class_for(category),
                "",
                precondition_for(category, requirement),
                data_for(req_id, category, requirement, note),
                expected_for(requirement),
                "",
            ]
        )

    apply_styles(ws)
    wb.save(SOURCE)
    print(f"saved={SOURCE}")
    print(f"sheet={sheet_name}")
    print(f"test_cases={len(requirements)}")


if __name__ == "__main__":
    main()
