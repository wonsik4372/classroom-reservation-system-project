from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "요구사항 관련기술"

# ── 데이터 ──────────────────────────────────────────────────────────────────
data = [
    # (카테고리, 연번, 요구사항, 관련기술)
    ("로그인\n(SFR-SL)", "SFR-SL001", "사용자의 정보를 입력받는다",
     "LoginGUI의 Swing 컴포넌트(JTextField, JPasswordField, JComboBox)를 통한 사용자 입력 수집 및 UserDTO.Request 객체로 캡슐화"),
    ("로그인\n(SFR-SL)", "SFR-SL002", "사용자의 정보를 검증한다",
     "LoginGUI 클라이언트 측 ID 형식 사전 검증 + LoginHandler에서 User.validateIdFormat() 및 User.verifyPassword()를 통한 이중 유효성 검증"),
    ("로그인\n(SFR-SL)", "SFR-SL003", "사용자의 권한을 인증한다",
     "Role enum 기반 역할 분리와 LoginHandler의 싱글톤(Initialization-on-demand Holder 패턴)에서 role 일치 여부 검증 후 UserDTO.Response에 인증 결과 반환"),
    ("로그인\n(SFR-SL)", "SFR-SL004", "사용자 정보 검증 결과에 대한 피드백 알림을 출력한다",
     "ResponseDTO의 success 플래그 및 message 필드를 SwingUtilities.invokeLater()로 EDT에서 JOptionPane 다이얼로그 출력"),
    ("로그인\n(SFR-SL)", "SFR-SL005", "사용자 권한 별 메인 대시보드를 로드한다",
     "LoginGUI에서 UserDTO.Response.role에 따라 AdminGUI / UserGUI / AssistantGUI로 분기하는 역할 기반 화면 라우팅"),
    ("로그인\n(SFR-SL)", "SFR-SL006", "사용자의 로그인 세션을 파기하고 초기화한다",
     "SessionManager.logout()으로 인메모리 세션 상태 초기화, NotificationController.stopPolling()으로 폴링 Swing Timer 정지"),

    ("사용자 관리\n(SFR-UM)", "SFR-UM001", "사용자 목록을 조회한다",
     "UserHandler.getUserList()로 UserCatalog(CopyOnWriteArrayList)에서 스레드 안전하게 조회, AdminGUI의 JTable에 렌더링"),
    ("사용자 관리\n(SFR-UM)", "SFR-UM002", "사용자를 추가한다",
     "AddUserGUI에서 UserDTO.Request로 입력 캡슐화, UserHandler.addUser()가 User 도메인 객체 생성 후 UserCatalog.addUser(synchronized) 및 UserFileManager로 JSON 영속화"),
    ("사용자 관리\n(SFR-UM)", "SFR-UM003", "사용자 추가에 대한 성공/실패를 알린다",
     "CRSystemClient.sendRequest()의 Consumer<ResponseDTO> 콜백으로 비동기 결과 수신, SwingUtilities.invokeLater()로 EDT 안전 피드백 출력"),
    ("사용자 관리\n(SFR-UM)", "SFR-UM004", "사용자를 삭제한다",
     "AdminGUI의 배치 선택 삭제 기능, UserHandler.deleteUser()가 UserCatalog.deleteUser(synchronized) 및 UserFileManager에서 순차 제거"),
    ("사용자 관리\n(SFR-UM)", "SFR-UM005", "사용자 삭제에 대한 성공/실패를 알린다",
     "순차적 비동기 콜백 체이닝(Consumer<ResponseDTO>)으로 각 삭제 결과를 개별 처리 후 JOptionPane으로 결과 알림"),

    ("강의실 조회\n(SFR-RV)", "SFR-RV001", "특정 강의실 정보를 검색한다",
     "ClassroomHandler.getMergedSchedule()이 Jackson ObjectMapper로 연도별 JSON을 역직렬화하여 ScheduleData(LinkedHashMap 기반) 반환"),
    ("강의실 조회\n(SFR-RV)", "SFR-RV002", "특정 강의실의 사용 가능한 빈 시간대를 추출한다",
     "Reservation DTO의 status 필드(0: 빈 시간)를 스트림 필터링으로 추출, TimeTableGUI / UserGUI에서 시각화"),

    ("강의실 관리\n(SFR-RM)", "SFR-RM001", "시스템의 강의실 정보를 관리한다",
     "ClassroomHandler(Initialization-on-demand Holder 싱글톤)의 updateClassroomInfo()로 ClassroomInfo 객체 갱신, synchronized saveBuildingFile()로 JSON 파일에 원자적 저장"),
    ("강의실 관리\n(SFR-RM)", "SFR-RM002", "EXCEL 형태의 시간표 파일을 수집한다",
     "ScheduleInitializer에서 Apache POI(WorkbookFactory)로 XLS/XLSX 파일 파싱, 연도·학기·건물·층·강의실 계층 구조 순회"),
    ("강의실 관리\n(SFR-RM)", "SFR-RM003", "수집한 시간표 파일의 데이터를 저장한다",
     "Apache POI로 파싱한 데이터를 Jackson ObjectMapper.writeValue()로 data/masterfile/YYYY.json에 직렬화 저장"),
    ("강의실 관리\n(SFR-RM)", "SFR-RM004", "강의실의 학기별 운영 시간을 설정한다",
     "TimetableHandler.updateRoomInfo(synchronized)로 Classroom DTO 내 ScheduleData의 운영 시간 필드 갱신 후 JSON 재저장"),

    ("예약 생성\n(SFR-CR)", "SFR-CR001", "예약 신청을 위한 예약자 기본 정보를 식별한다",
     "SessionManager.getInstance()로 현재 로그인 사용자의 UserDTO.Response(userId, userName, role)를 조회하여 ReservationDTO.Request에 주입"),
    ("예약 생성\n(SFR-CR)", "SFR-CR002", "예약 신청 데이터의 시간 단위 유효성을 검증한다",
     "ReservationHandler.addReservation()에서 LocalDate 파싱 및 period 범위 검증, ClassroomHandler와의 교차 검증으로 충돌 감지"),
    ("예약 생성\n(SFR-CR)", "SFR-CR003", "교수 권한 사용자의 예약을 즉시 승인한다",
     "ReservationHandler에서 Role.PROFESSOR 분기 시 ReservationDTO.Status.APPROVED로 직접 설정 후 ReservationCatalog 저장"),
    ("예약 생성\n(SFR-CR)", "SFR-CR004", "교수 예약 시간대와 겹치는 학생 예약을 자동 취소한다",
     "ReservationHandler의 교수 예약 승인 시 ReservationCatalog 전체 순회, 동일 강의실·시간대 PENDING 학생 예약을 REJECTED 처리 및 NotificationHandler.notifyRejected(\"교수 보강\") 호출"),
    ("예약 생성\n(SFR-CR)", "SFR-CR005", "교수 권한의 최대 예약 가능 시간을 제한한다",
     "ReservationHandler에서 교수의 일별 예약 period 합산 후 최대 3교시 초과 시 거부 로직"),
    ("예약 생성\n(SFR-CR)", "SFR-CR006", "학생 권한의 예약을 대기 상태로 생성한다",
     "Role.STUDENT 분기 시 ReservationDTO.Status.PENDING으로 설정, ReservationCatalog 및 ReservationFileManager에 영속화"),
    ("예약 생성\n(SFR-CR)", "SFR-CR007", "학생 권한의 최소 예약 신청 시점을 검증한다",
     "ReservationHandler에서 LocalDate.now().plusDays(1)과 예약 요청 날짜 비교, 당일 예약 시도 차단"),
    ("예약 생성\n(SFR-CR)", "SFR-CR008", "학생 권한의 최대 예약 가능 시간을 제한한다",
     "ReservationHandler에서 학생의 일별 예약 period 합산 후 최대 2교시 초과 시 거부 로직"),
    ("예약 생성\n(SFR-CR)", "SFR-CR009", "학생 예약의 강의실 수용 인원 부합 여부를 검증한다",
     "ClassroomInfo.capacity와 ReservationDTO.Request.partnerCount를 비교하여 수용 인원 50% 초과 시 거부"),

    ("예약 관리\n(SFR-MR)", "SFR-MR001", "대기 상태의 예약을 단일 또는 다중 선택하여 승인한다",
     "AssistantGUI의 JTable 다중 선택(getSelectedRows()) + ReservationController.approveReservation()의 순차적 비동기 콜백 체이닝"),
    ("예약 관리\n(SFR-MR)", "SFR-MR002", "대기 상태의 예약을 선택하여 거부한다",
     "AssistantGUI에서 선택된 ReservationDTO.Response를 ReservationController.rejectReservation()으로 전달"),
    ("예약 관리\n(SFR-MR)", "SFR-MR003", "거부한 예약의 거부 사유를 입력한다",
     "JOptionPane.showInputDialog()로 사유 입력 후 ReservationDTO.Request.rejectReason 필드에 설정하여 RequestDTO 페이로드로 직렬화 전송"),
    ("예약 관리\n(SFR-MR)", "SFR-MR004", "선택된 예약의 상태를 업데이트한다",
     "ReservationHandler에서 ReservationCatalog 인메모리 상태 변경 + ReservationFileManager.saveAll()로 JSON 파일 동기 저장"),
    ("예약 관리\n(SFR-MR)", "SFR-MR005", "예약 승인 결과를 예약자에게 안내한다",
     "NotificationHandler.notifyApproved()로 NotificationCatalog(ConcurrentHashMap + CopyOnWriteArrayList)에 NotificationDTO(Type.APPROVED) 저장, 클라이언트 폴링으로 수신"),
    ("예약 관리\n(SFR-MR)", "SFR-MR006", "예약 거부 결과와 사유를 예약자에게 안내한다",
     "NotificationHandler.notifyRejected(rejectReason)으로 NotificationDTO(Type.REJECTED, rejectReason) 저장, NotificationController 5초 Swing Timer 폴링으로 수신 후 다이얼로그 표시"),
    ("예약 관리\n(SFR-MR)", "SFR-MR007", "본인의 예약 내역을 직접 취소한다",
     "UserGUI의 내 예약 탭에서 ReservationController.cancelReservation()으로 SessionManager의 userId 기반 소유권 검증 후 PENDING 예약 삭제"),

    ("데이터 관리\n(SFR-DM)", "SFR-DM001", "백업된 파일을 롤백한다",
     "FileManager<T> 인터페이스 구현체를 통한 백업 JSON 파일 재로드 및 saveAll()로 마스터 파일 덮어쓰기"),
    ("데이터 관리\n(SFR-DM)", "SFR-DM002", "데이터 복구 진행 전 작업의 위험성을 경고한다",
     "JOptionPane.showConfirmDialog()를 통한 사용자 확인 대화상자 표시"),
    ("데이터 관리\n(SFR-DM)", "SFR-DM003", "시스템의 마스터 데이터를 백업한다",
     "UserFileManager / ReservationFileManager의 Gson prettyPrinting 직렬화로 타임스탬프 파일명의 백업 JSON 생성"),
    ("데이터 관리\n(SFR-DM)", "SFR-DM004", "백업 파일의 시간 정보를 표기한다",
     "LocalDateTime.now()를 파일명 패턴(YYYY-MM-DD_HH-mm-ss)에 포맷하여 백업 파일명에 포함"),
    ("데이터 관리\n(SFR-DM)", "SFR-DM005", "프로그램 종료 시점의 변경 사항을 보존한다",
     "CRSystemServer의 JVM Shutdown Hook(Runtime.getRuntime().addShutdownHook())에서 UserFileManager.saveAll() / ReservationFileManager.saveAll() 호출"),
    ("데이터 관리\n(SFR-DM)", "SFR-DM006", "실행 최초의 필요한 데이터를 자동 생성한다",
     "ScheduleInitializer가 서버 기동 시 마스터 JSON 파일 부재 확인 후 Excel 파일로부터 초기 데이터 자동 생성"),

    ("네트워크/세션\n(SFR-NM)", "SFR-NM001", "시스템의 실시간 접속 현황을 추적한다",
     "SessionRegistry(ConcurrentHashMap + AtomicInteger 세션 ID 채번)로 연결된 ClientHandler 목록과 접속 시각 추적"),
    ("네트워크/세션\n(SFR-NM)", "SFR-NM002", "세션의 비정상 접속을 감지한다",
     "ClientHandler의 ScheduledExecutorService 워치독이 AtomicInteger missedHeartbeats를 주기적으로 증가시켜 3회(MAX_MISSED_HEARTBEATS) 초과 시 비정상으로 판단"),
    ("네트워크/세션\n(SFR-NM)", "SFR-NM003", "비정상 종료된 세션의 시스템 자원을 회수한다",
     "ClientHandler.closeSession()에서 소켓 스트림 닫기, SessionRegistry.unregister(sessionId), ScheduledExecutorService.shutdown() 순차 실행으로 자원 반납"),
    ("네트워크/세션\n(SFR-NM)", "SFR-NM004", "수신 요청 메시지의 데이터 규격을 검증한다",
     "MainController의 HashMap<String, Command> 명령 라우팅에서 알 수 없는 command 수신 시 ResponseDTO(success=false) 반환, RequestDTO 역직렬화 실패 시 ClassNotFoundException / IOException 처리"),
    ("네트워크/세션\n(SFR-NM)", "SFR-NM005", "불량 데이터의 예외 처리를 수행한다",
     "ClientHandler의 try-catch 블록에서 EOFException(정상 종료), SocketException(네트워크 단절), Exception(기타) 분기 처리 후 closeSession() 호출"),
]

# ── 색상 팔레트 ───────────────────────────────────────────────────────────────
CAT_COLORS = {
    "로그인\n(SFR-SL)":       ("1F4E79", "D6E4F0"),
    "사용자 관리\n(SFR-UM)":   ("375623", "E2EFDA"),
    "강의실 조회\n(SFR-RV)":   ("7B2D8B", "F3E5F5"),
    "강의실 관리\n(SFR-RM)":   ("8B4513", "FFF3E0"),
    "예약 생성\n(SFR-CR)":     ("C00000", "FDECEA"),
    "예약 관리\n(SFR-MR)":     ("833C00", "FFF8E1"),
    "데이터 관리\n(SFR-DM)":   ("1B4F72", "EAF4FB"),
    "네트워크/세션\n(SFR-NM)": ("4A235A", "F5EEF8"),
}

thin = Side(style="thin", color="AAAAAA")
medium = Side(style="medium", color="555555")
border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
border_header = Border(left=medium, right=medium, top=medium, bottom=medium)

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

# ── 헤더 ─────────────────────────────────────────────────────────────────────
headers = ["카테고리", "연번", "요구사항", "관련 기술"]
col_widths = [16, 13, 32, 90]

for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
    cell = ws.cell(row=1, column=ci, value=h)
    cell.font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)
    cell.fill = make_fill("2C3E50")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border_header
    ws.column_dimensions[get_column_letter(ci)].width = w

ws.row_dimensions[1].height = 28

# ── 데이터 행 삽입 (카테고리 병합 준비) ──────────────────────────────────────
row = 2
cat_group: dict[str, list[int]] = {}

for cat, sfr_id, req, tech in data:
    dark, light = CAT_COLORS[cat]

    # 카테고리 열
    c1 = ws.cell(row=row, column=1, value=cat)
    c1.font = Font(name="맑은 고딕", bold=True, color=dark, size=10)
    c1.fill = make_fill(light)
    c1.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c1.border = border_thin

    # 연번 열
    c2 = ws.cell(row=row, column=2, value=sfr_id)
    c2.font = Font(name="맑은 고딕", bold=True, color=dark, size=10)
    c2.fill = make_fill(light)
    c2.alignment = Alignment(horizontal="center", vertical="center")
    c2.border = border_thin

    # 요구사항 열
    c3 = ws.cell(row=row, column=3, value=req)
    c3.font = Font(name="맑은 고딕", size=10)
    c3.fill = make_fill("FAFAFA")
    c3.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c3.border = border_thin

    # 관련기술 열
    c4 = ws.cell(row=row, column=4, value=tech)
    c4.font = Font(name="맑은 고딕", size=10)
    c4.fill = make_fill("FFFFFF")
    c4.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c4.border = border_thin

    ws.row_dimensions[row].height = 42

    cat_group.setdefault(cat, []).append(row)
    row += 1

# ── 카테고리 셀 병합 ──────────────────────────────────────────────────────────
for cat, rows in cat_group.items():
    if len(rows) > 1:
        ws.merge_cells(start_row=rows[0], start_column=1,
                       end_row=rows[-1],  end_column=1)
        merged = ws.cell(row=rows[0], column=1)
        dark, light = CAT_COLORS[cat]
        merged.font = Font(name="맑은 고딕", bold=True, color=dark, size=10)
        merged.fill = make_fill(light)
        merged.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        merged.border = Border(left=medium, right=thin, top=medium, bottom=medium)

# ── 틀 고정 & 자동 필터 ───────────────────────────────────────────────────────
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:D{row - 1}"

# ── 시트 여백 & 인쇄 설정 ─────────────────────────────────────────────────────
ws.sheet_view.showGridLines = True
ws.print_title_rows = "1:1"

out = r"C:\Users\KANG\Desktop\test-clone\classroom-reservation-system-project\docs\SFR_관련기술.xlsx"
wb.save(out)
print(f"저장 완료: {out}")
