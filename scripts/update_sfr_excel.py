import openpyxl
import copy

src = r'C:\Users\KANG\Desktop\sw-project-v1\Requirement\SE-0.SW요구사항-강의실예약시스템-1조.xlsx'
dst = r'C:\Users\KANG\Desktop\sw-project-v1\Requirement\SE-0.SW요구사항-강의실예약시스템-1조.xlsx'

wb = openpyxl.load_workbook(src)
ws = wb['SRS Test Cases']

# TC-47 ~ TC-51 업데이트 정보
# (TC ID, 새 테스팅결과, 새 테스트메서드)
updates = {
    'TC-47': (
        'PASS',
        'connectTime_isSetOnCreation\nrun_registersHandlerInSessionRegistry\nloadProperties_loadsPortFromPropertiesFile\nloadProperties_usesDefaultPortWhenPropertiesFileMissing'
    ),
    'TC-48': (
        'PASS',
        'onHeartbeatReceived_resetsMissedHeartbeatCount\nrun_respondsToHeartbeatWithPong'
    ),
    'TC-49': (
        'PASS',
        'run_unregistersSessionOnDisconnect'
    ),
    'TC-50': (
        'PASS',
        'run_returnsFailResponseForUnknownCommand'
    ),
    'TC-51': (
        'PASS',
        'run_sendsErrorResponseForNonRequestDTOObject'
    ),
}

# 헤더 행에서 컬럼 인덱스 찾기
header_row = None
tc_col = result_col = method_col = None

for row in ws.iter_rows():
    for cell in row:
        if cell.value == 'ID':
            header_row = cell.row
            tc_col = cell.column
        if cell.value == '테스팅 결과':
            result_col = cell.column
        if cell.value == '테스트 메서드':
            method_col = cell.column
    if header_row:
        break

print(f"헤더 행: {header_row}, TC컬럼: {tc_col}, 결과컬럼: {result_col}, 메서드컬럼: {method_col}")

# 데이터 행 순회하여 해당 TC 업데이트
for row in ws.iter_rows(min_row=header_row + 1):
    tc_cell = row[tc_col - 1]
    if tc_cell.value in updates:
        new_result, new_methods = updates[tc_cell.value]
        result_cell = row[result_col - 1]
        method_cell = row[method_col - 1]

        result_cell.value = new_result
        method_cell.value = new_methods

        print(f"  Updated {tc_cell.value}: 결과={new_result}")

wb.save(dst)
print(f"\n저장 완료: {dst}")
